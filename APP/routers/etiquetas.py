from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from typing import Optional
import io
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import barcode
from barcode.writer import ImageWriter
from PIL import Image as PILImage, ImageDraw, ImageFont
from APP.db import engine
from datetime import date

router = APIRouter(prefix="/etiquetas", tags=["etiquetas"])

COLS_PER_ROW = 4
# Cell pixel dimensions — must stay in sync with COL_WIDTH / ROW_HEIGHT below.
# COL_WIDTH=46 chars → ~330px at 96dpi.  ROW_HEIGHT=90pt → 120px at 96dpi.
CELL_W_PX   = 330   # px — barcode fills this width exactly
CELL_H_PX   = 120   # px — total cell height
TEXT_H_PX   = 15    # px — bottom strip reserved for SKU text
SKU_FONT_PT = 13    # fixed font size; text never scales with barcode width
MODULE_WIDTH  = 1.2   # mm per bar module — drives initial render resolution
MODULE_HEIGHT = 14.0  # mm bar height (overridden when fitting to cell)
# Excel column width (chars) and row height (pt)
COL_WIDTH  = 46
ROW_HEIGHT = 90   # pt


class BarcodeSheetRequest(BaseModel):
    categoria_id: Optional[int] = None
    include_children: bool = True
    skus: list[str] = []
    format: str = "CODE128"


def _get_products(req: BarcodeSheetRequest) -> list[str]:
    """Return list of SKUs matching the request."""
    with engine.connect() as conn:
        if req.skus:
            rows = conn.execute(
                text("SELECT sku FROM productos WHERE sku = ANY(:skus) AND is_active = true ORDER BY sku"),
                {"skus": req.skus},
            ).fetchall()
            return [r[0] for r in rows]

        if req.categoria_id is None:
            return []

        if req.include_children:
            sql = text("""
                WITH RECURSIVE cat_tree AS (
                    SELECT id FROM categoria WHERE id = :cid
                    UNION ALL
                    SELECT c.id FROM categoria c
                    JOIN cat_tree ct ON c.parent_id = ct.id
                )
                SELECT sku FROM productos
                WHERE categoria_id IN (SELECT id FROM cat_tree)
                  AND is_active = true
                  AND catalog_visible = true
                ORDER BY sku
            """)
        else:
            sql = text("""
                SELECT sku FROM productos
                WHERE categoria_id = :cid
                  AND is_active = true
                  AND catalog_visible = true
                ORDER BY sku
            """)

        rows = conn.execute(sql, {"cid": req.categoria_id}).fetchall()
        return [r[0] for r in rows]


def _make_barcode_image(sku: str) -> tuple[io.BytesIO, int, int]:
    """Returns (image_buf, CELL_W_PX, CELL_H_PX).

    Barcode bars fill the full cell width exactly.
    SKU text is rendered at a fixed font size in a bottom strip —
    it never scales with bar width so it stays legible for any SKU length.
    """
    writer = ImageWriter()
    bc = barcode.get("code128", sku, writer=writer)
    bc_buf = io.BytesIO()
    bc.write(bc_buf, options={
        "module_width":  MODULE_WIDTH,
        "module_height": MODULE_HEIGHT,
        "quiet_zone":    0.5,
        "write_text":    False,
        "background":    "white",
        "foreground":    "black",
    })
    bc_buf.seek(0)

    bars_h = CELL_H_PX - TEXT_H_PX
    bc_img = PILImage.open(bc_buf).convert("RGB")
    # Scale to exact cell width; stretch height to fill bars area
    bc_img = bc_img.resize((CELL_W_PX, bars_h), PILImage.LANCZOS)

    canvas = PILImage.new("RGB", (CELL_W_PX, CELL_H_PX), "white")
    canvas.paste(bc_img, (0, 0))

    # SKU text — fixed size, centered in bottom strip
    draw = ImageDraw.Draw(canvas)
    try:
        font = ImageFont.truetype("cour.ttf", SKU_FONT_PT)
    except OSError:
        font = ImageFont.load_default()

    bbox = draw.textbbox((0, 0), sku, font=font)
    tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    tx = max(0, (CELL_W_PX - tw) // 2)
    ty = bars_h + 1
    draw.text((tx, ty), sku, fill="black", font=font)

    out = io.BytesIO()
    canvas.save(out, format="PNG", dpi=(150, 150))
    out.seek(0)
    return out, CELL_W_PX, CELL_H_PX


def _build_workbook(skus: list[str], sheet_title: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    # Set column widths
    for col in range(1, COLS_PER_ROW + 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTH

    excel_row = 1
    col_idx   = 0  # 0-based within current row

    for sku in skus:
        if col_idx == 0:
            # Set row height for barcode row
            ws.row_dimensions[excel_row].height = ROW_HEIGHT

        col_letter = get_column_letter(col_idx + 1)
        cell_addr  = f"{col_letter}{excel_row}"

        # Write SKU as cell value (visible if image is missing, useful for copy)
        cell = ws[cell_addr]
        cell.value     = sku
        cell.font      = Font(name="Courier New", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)

        # Insert barcode image anchored to this cell
        try:
            img_buf, img_w, img_h = _make_barcode_image(sku)
            xl_img        = XLImage(img_buf)
            xl_img.anchor = cell_addr
            xl_img.width  = img_w
            xl_img.height = img_h
            ws.add_image(xl_img)
        except Exception:
            # If barcode generation fails for this SKU, leave text only
            pass

        col_idx += 1
        if col_idx >= COLS_PER_ROW:
            col_idx   = 0
            excel_row += 1

    # If last row is partial, still set its height
    if col_idx != 0:
        ws.row_dimensions[excel_row].height = ROW_HEIGHT

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("/barcodes/sheet")
def generate_barcode_sheet(req: BarcodeSheetRequest):
    skus = _get_products(req)

    if not skus:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="No se encontraron productos para los criterios indicados.")

    # Sheet title: use categoria name if available, else "Codigos"
    sheet_title = "Codigos"
    if req.categoria_id:
        with engine.connect() as conn:
            row = conn.execute(
                text("SELECT name FROM categoria WHERE id = :id"),
                {"id": req.categoria_id},
            ).fetchone()
            if row:
                sheet_title = row[0][:31]

    filename_label = sheet_title.replace(" ", "_").lower()
    filename = f"codigos_{filename_label}_{date.today().isoformat()}.xlsx"

    workbook_buf = _build_workbook(skus, sheet_title)

    return StreamingResponse(
        workbook_buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
